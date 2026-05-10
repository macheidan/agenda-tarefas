"""Extrai (nome, handle) da aba 'Lista 2026' de um XLSX recebido como JSON+base64."""
import base64
import json
import re
import sys
from pathlib import Path

import openpyxl

SRC_JSON = Path(
    r"C:\Users\PICHAU\.claude\projects\C--claude-project-Pizzarias-intranet-pizzarias"
    r"\deda7978-553b-491f-84dc-5aea1e34d553\tool-results"
    r"\mcp-claude_ai_Google_Drive-download_file_content-1778387249369.txt"
)
OUT_XLSX = Path(r"C:\claude_project\Pizzarias\intranet-pizzarias\scripts\lista-2026.xlsx")
OUT_JSON = Path(r"C:\claude_project\Pizzarias\intranet-pizzarias\src\data\handles-import.json")


def decode_xlsx() -> None:
    raw = json.loads(SRC_JSON.read_text(encoding="utf-8"))
    content_b64 = raw["content"]
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    OUT_XLSX.write_bytes(base64.b64decode(content_b64))
    print(f"[ok] xlsx escrito em {OUT_XLSX} ({OUT_XLSX.stat().st_size} bytes)")


def find_sheet(wb) -> str:
    target = "lista 2026"
    for name in wb.sheetnames:
        if target in name.strip().lower():
            return name
    # fallback: primeira aba
    print(f"[warn] aba 'Lista 2026' nao encontrada. Abas: {wb.sheetnames}", file=sys.stderr)
    return wb.sheetnames[0]


def url_to_handle(url: str) -> str:
    if not url:
        return ""
    url = url.strip()
    # Instagram
    m = re.match(r"https?://(?:www\.)?instagram\.com/([^/?#]+)/?", url, re.IGNORECASE)
    if m:
        return m.group(1)
    # Generico: ultimo segmento path
    try:
        from urllib.parse import urlparse

        path = urlparse(url).path.strip("/")
        if path:
            return path.split("/")[-1]
    except Exception:
        pass
    return ""


def extract() -> None:
    wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
    sheet_name = find_sheet(wb)
    ws = wb[sheet_name]
    print(f"[ok] usando aba: {sheet_name!r} (max_row={ws.max_row})")

    # Detectar header — procurar linha que tenha 'NOME' na coluna B
    header_row = 1
    for row_idx in range(1, min(10, ws.max_row + 1)):
        val = ws.cell(row=row_idx, column=2).value
        if isinstance(val, str) and "nome" in val.strip().lower():
            header_row = row_idx
            break
    print(f"[ok] header_row={header_row}")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=2)
        nome = cell.value
        if nome is None or (isinstance(nome, str) and not nome.strip()):
            continue
        nome = str(nome).strip()
        handle = ""
        if cell.hyperlink and cell.hyperlink.target:
            handle = url_to_handle(cell.hyperlink.target)
        rows.append({"nome": nome, "handle": handle})

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    com_handle = sum(1 for r in rows if r["handle"])
    print(f"[ok] total={len(rows)} com_handle={com_handle}")
    print("[exemplos]")
    for r in rows[:5]:
        print(f"  - {r}")
    print(f"[ok] gravado em {OUT_JSON}")


if __name__ == "__main__":
    decode_xlsx()
    extract()
